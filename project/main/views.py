import csv
from tempfile import NamedTemporaryFile
import datetime

from openpyxl import Workbook, load_workbook
from django.http import HttpResponse
from rest_framework import viewsets
from rest_framework import permissions
from rest_framework import exceptions
from rest_framework import status
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FileUploadParser, JSONParser
from rest_framework.decorators import action

from main.serializers import VehicleSerializer
from main.models import Vehicle
from main.filters import VehicleFilter


class VehicleViewSet(viewsets.ModelViewSet):
    """
    A simple ViewSet for viewing and editing the Vehicles.
    """

    queryset = Vehicle.objects.all()
    serializer_class = VehicleSerializer
    permission_classes = [permissions.IsAuthenticated]
    filterset_class = VehicleFilter
    parser_classes = [JSONParser, MultiPartParser, FileUploadParser]

    def list(self, request, *args, **kwargs):
        download_query = request.query_params.get("download")
        if download_query == "csv":
            return self.download_csv(request, *args, **kwargs)
        if download_query == "xlsx":
            return self.download_xlsx(request, *args, **kwargs)

        return super().list(request, *args, **kwargs)

    def download_csv(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())  # as in super().list()

        response = HttpResponse(
            headers={
                "Content-Type": "text/csv",
                "Content-Disposition": 'attachment; filename="vehicles_list.csv"',
            }
        )
        fieldnames = [field.name for field in Vehicle._meta.get_fields()]
        writer = csv.DictWriter(response, fieldnames=fieldnames)
        writer.writeheader()
        for row in queryset.values():
            writer.writerow(row)

        return response

    def download_xlsx(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())  # as in super().list()

        wb = Workbook()
        ws = wb.active
        ws.title = "Vehicles"

        fieldnames = [field.name for field in Vehicle._meta.get_fields()]
        ws.append(fieldnames)
        for row in queryset.values():
            ws.append(list(row.values()))

        with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()

        response = HttpResponse(
            stream,
            headers={
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "Content-Disposition": 'attachment; filename="vehicles_list.xlsx"',
            },
        )
        return response

    @action(detail=False, methods=["post"])
    def from_csv(self, request):
        file = request.data.get("file")
        if file is None:
            raise exceptions.ParseError("Файл не был получен")
        rows = [row.decode() for row in file]
        reader = csv.DictReader(rows)
        good = []  # list of created Vehicles
        bad = []  # list of bad data
        for data in reader:
            serializer = VehicleSerializer(data=data, context={"request": request})
            res = serializer.is_valid()
            if res:
                vehicle = serializer.save()
                good.append(vehicle)
            else:
                bad.append({"data": data, "errors": serializer.errors})
        return_data = {
            "created": VehicleSerializer(
                good, many=True, context={"request": request}
            ).data,
            "bad_data": bad,
        }

        return Response(return_data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=["post"])
    def from_xlsx(self, request):
        file = request.data.get("file")
        if file is None:
            raise exceptions.ParseError("Файл не был получен")
        with NamedTemporaryFile(suffix=".xlsx") as tmp:
            tmp.write(file.read())
            tmp.seek(0)
            wb = load_workbook(tmp.name)
        ws = wb.active
        fieldnames = [str(cell.value) for cell in ws["1:1"]]
        good = []  # list of created Vehicles
        bad = []  # list of bad data
        for row in ws.iter_rows(min_row=2):
            data = {}
            for fieldname, cell in zip(fieldnames, row):
                if fieldname == "vehicle_registration_date" and isinstance(
                    cell.value, datetime.datetime
                ):
                    data[fieldname] = cell.value.strftime("%Y-%m-%d")
                else:
                    data[fieldname] = cell.value
            serializer = VehicleSerializer(data=data, context={"request": request})
            res = serializer.is_valid()
            if res:
                vehicle = serializer.save()
                good.append(vehicle)
            else:
                bad.append({"data": data, "errors": serializer.errors})
        return_data = {
            "created": VehicleSerializer(
                good, many=True, context={"request": request}
            ).data,
            "bad_data": bad,
        }

        return Response(return_data, status=status.HTTP_201_CREATED)
